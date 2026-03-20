"""
Excel Exporter for Pre-Open Scanner
Creates Excel file with 7 sheets (1 per minute from 9:00-9:07 AM)

Run at 8:55 AM before market opens.
"""

import os
import time
from datetime import datetime, time as dt_time, timedelta
from dotenv import load_dotenv
import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

load_dotenv()

# Backend API
BACKEND_URL = "https://pre-open-scanner-production.up.railway.app"

# Output file
OUTPUT_FILE = "preopen_data.xlsx"

# Data storage: {minute: [list of tick data]}
minute_data = {
    "09:00": [],
    "09:01": [],
    "09:02": [],
    "09:03": [],
    "09:04": [],
    "09:05": [],
    "09:06": []
}


def get_stocks_data():
    """Fetch all stocks from backend."""
    try:
        response = requests.get(
            f"{BACKEND_URL}/api/stocks",
            params={"limit": 500},
            timeout=5
        )
        if response.status_code == 200:
            data = response.json()
            return data.get("stocks", [])
        return []
    except Exception as e:
        print(f"Error fetching data: {e}")
        return []


def format_stock_data(stocks):
    """Convert stock data to DataFrame."""
    rows = []
    for stock in stocks:
        symbol = stock.get("symbol", "").replace("NSE:", "").replace("-EQ", "")
        rows.append({
            "Symbol": symbol,
            "LTP": stock.get("ltp", 0),
            "Change": stock.get("change", 0),
            "Change %": stock.get("change_pct", 0),
            "Volume": stock.get("volume", 0),
            "Buy Qty": stock.get("buy_qty", 0),
            "Sell Qty": stock.get("sell_qty", 0),
            "High": stock.get("high", 0),
            "Low": stock.get("low", 0),
            "Open": stock.get("open", 0),
            "Prev Close": stock.get("prev_close", 0),
            "Timestamp": stock.get("timestamp", "")
        })
    return pd.DataFrame(rows)


def create_excel_file():
    """Create Excel file with 7 sheets."""
    print("\n" + "="*60)
    print("CREATING EXCEL FILE")
    print("="*60)
    
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Create 7 sheets
    for minute_key, data_list in minute_data.items():
        print(f"\nSheet: {minute_key} - {len(data_list)} records")
        
        if not data_list:
            print(f"  ⚠️ No data for {minute_key}")
            continue
        
        # Take the last snapshot for this minute
        stocks = data_list[-1] if data_list else []
        
        if not stocks:
            continue
        
        # Convert to DataFrame
        df = format_stock_data(stocks)
        
        # Create sheet
        ws = wb.create_sheet(title=minute_key)
        
        # Add header row with styling
        header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                # Header styling
                if r_idx == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")
                
                # Number formatting
                if r_idx > 1:
                    if c_idx in [2, 3, 8, 9, 10, 11]:  # LTP, Change, High, Low, Open, Prev Close
                        cell.number_format = '₹#,##0.00'
                    elif c_idx == 4:  # Change %
                        cell.number_format = '0.00"%"'
                    elif c_idx in [5, 6, 7]:  # Volume, Buy Qty, Sell Qty
                        cell.number_format = '#,##0'
                    
                    # Color coding for Change %
                    if c_idx == 4:
                        if value and float(value) > 0:
                            cell.font = Font(color="008000")  # Green
                        elif value and float(value) < 0:
                            cell.font = Font(color="FF0000")  # Red
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        print(f"  ✅ Sheet created: {len(df)} stocks")
    
    # Save file
    wb.save(OUTPUT_FILE)
    print(f"\n✅ Excel file saved: {OUTPUT_FILE}")
    print("="*60)


def main():
    """Main loop - capture data every 10 seconds, save by minute."""
    print("="*60)
    print("EXCEL EXPORTER - 7 SHEETS (9:00-9:07 AM)")
    print("="*60)
    
    # Wait until 9:00 AM
    now = datetime.now()
    target_start = now.replace(hour=9, minute=0, second=0, microsecond=0)
    target_end = now.replace(hour=9, minute=7, second=59, microsecond=0)
    
    if now < target_start:
        wait_seconds = (target_start - now).total_seconds()
        print(f"\n⏳ Waiting until 9:00 AM ({wait_seconds:.0f} seconds)...")
        time.sleep(wait_seconds)
    
    print("\n📊 CAPTURING DATA (9:00-9:07 AM)...\n")
    
    last_minute = None
    
    while datetime.now() <= target_end:
        current_time = datetime.now()
        current_minute = current_time.strftime("%H:%M")
        
        # Capture data
        stocks = get_stocks_data()
        
        if stocks:
            # Determine which minute bucket (9:00, 9:01, etc.)
            minute_key = current_minute[:5]  # "09:00", "09:01", etc.
            
            if minute_key in minute_data:
                minute_data[minute_key].append(stocks)
                
                if minute_key != last_minute:
                    print(f"\n[{current_time.strftime('%H:%M:%S')}] Now capturing: {minute_key}")
                    last_minute = minute_key
                else:
                    print(f"[{current_time.strftime('%H:%M:%S')}] Tick captured ({len(stocks)} stocks)")
        
        time.sleep(10)  # Capture every 10 seconds
    
    print("\n\n" + "="*60)
    print("CAPTURE COMPLETE!")
    print("="*60)
    
    # Show summary
    for minute, data in minute_data.items():
        print(f"{minute}: {len(data)} snapshots captured")
    
    # Create Excel file
    create_excel_file()
    
    print(f"\n✅ DONE! File ready: {OUTPUT_FILE}")
    print("📧 You can now send this to your CEO or Telegram!")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n🛑 Export stopped by user")
        print("Creating Excel from captured data...")
        create_excel_file()
