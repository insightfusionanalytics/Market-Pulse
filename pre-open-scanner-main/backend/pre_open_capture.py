"""
NSE PRE-OPEN LIVE CAPTURE — v5 FINAL (MERGED BEST OF V2 + V4)
==============================================================
HOW TO RUN:
    cd backend
    python preopen_capture_v5.py

You can run this ANY TIME before 8:58 AM.
It will wait, then auto-connect at 8:58.
It will auto-stop at 9:20. No manual action needed.

OUTPUTS (saved to backend/outputs/):
    01_raw_ticks_YYYYMMDD.csv       — Every single tick
    02_per_second_YYYYMMDD.csv      — One row per symbol per second
    03_phase_summary_YYYYMMDD.csv   — Snapshot at each phase trigger
    04_analysis_YYYYMMDD.xlsx       — Excel: 5 phase sheets + summary

TELEGRAM: 8 snapshots + instant IEP-change alerts per stock
"""

import csv
import os
import threading
import time
from datetime import datetime

import requests
from dotenv import load_dotenv
from fyers_apiv3.FyersWebsocket import data_ws

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False
    print("⚠️  openpyxl not installed → run: pip install openpyxl")

load_dotenv()

# ═══════════════════════════════════════════════════════
#  CONFIG
# ═══════════════════════════════════════════════════════
APP_ID       = os.getenv("FYERS_APP_ID",       "OJH52PHNOP-100")
ACCESS_TOKEN = os.getenv("FYERS_ACCESS_TOKEN",  "")
TG_TOKEN     = os.getenv("TELEGRAM_BOT_TOKEN",  "8449739735:AAHe4sFKkQ81nDyvImu9WMMTlMr5DwjUYfo")
TG_CHAT_ID   = os.getenv("TELEGRAM_CHAT_ID",    "1004112832")
FULL_TOKEN   = f"{APP_ID}:{ACCESS_TOKEN}"

SYMBOLS = [
    "NSE:RELIANCE-EQ",
    "NSE:MAZDOCK-EQ",
    "NSE:HINDCOPPER-EQ",
    "NSE:ETERNAL-EQ",
    "NSE:RVNL-EQ",
]

START_TIME = "08:58:00"   # wait until this before connecting
STOP_TIME  = "09:20:00"   # auto-stop at this time
DATE_STR   = datetime.now().strftime("%Y%m%d")

# ═══════════════════════════════════════════════════════
#  PHASE BOUNDARIES  (V2's accurate 9:07:30)
# ═══════════════════════════════════════════════════════
def get_phase(hms: str) -> str:
    if   hms < "09:00:00": return "PRE_BASELINE"
    elif hms < "09:07:30": return "PHASE1_ORDER_COLLECTION"
    elif hms < "09:12:00": return "PHASE2_MATCHING"
    elif hms < "09:15:00": return "PHASE3_BUFFER"
    else:                  return "REGULAR_MARKET"

# ═══════════════════════════════════════════════════════
#  TELEGRAM SNAPSHOTS  (V4's 8 triggers)
# ═══════════════════════════════════════════════════════
SNAPSHOTS = {
    "09:00:15": "🟢 PHASE 1 START — Order Collection Begins",
    "09:04:00": "📊 PHASE 1 MID — 4 Minute Mark",
    "09:07:30": "🔴 PHASE 1 END — Order Window Closing",
    "09:08:30": "⚡ PHASE 2 — Auction Matching",
    "09:12:00": "🔄 PHASE 3 — Buffer / Transition",
    "09:14:45": "🚀 T-15s — Just Before Market Open",
    "09:15:15": "🏁 REGULAR MARKET OPEN",
    "09:20:00": "✅ AUTO-STOP — Session Complete",
}
sent_snaps  : set = set()
iep_alerted : set = set()   # per-stock IEP first-move alert

# ═══════════════════════════════════════════════════════
#  STATE  (all writes behind lock — V4's thread safety)
# ═══════════════════════════════════════════════════════
lock             = threading.Lock()
stop_event       = threading.Event()

latest           : dict = {}   # symbol → latest full tick dict
tick_count       : dict = {}   # symbol → int
all_keys         : set  = set()
ltp_history      : dict = {}   # symbol → [ltp values]
one_sec_log      : dict = {}   # "HH:MM:SS" → {symbol → dict}
written_sec_keys : set  = set()

# ═══════════════════════════════════════════════════════
#  FILES
# ═══════════════════════════════════════════════════════
os.makedirs("outputs", exist_ok=True)
RAW_PATH   = f"outputs/01_raw_ticks_{DATE_STR}.csv"
SEC_PATH   = f"outputs/02_per_second_{DATE_STR}.csv"
PHASE_PATH = f"outputs/03_phase_summary_{DATE_STR}.csv"
XLS_PATH   = f"outputs/04_analysis_{DATE_STR}.xlsx"

RAW_FILE   = open(RAW_PATH,   "w", newline="", encoding="utf-8", buffering=1)
SEC_FILE   = open(SEC_PATH,   "w", newline="", encoding="utf-8", buffering=1)
PHASE_FILE = open(PHASE_PATH, "w", newline="", encoding="utf-8", buffering=1)
raw_writer = sec_writer = phase_writer = None

# ═══════════════════════════════════════════════════════
#  HELPERS
# ═══════════════════════════════════════════════════════
def short(sym: str) -> str:
    return sym.split(":")[1].replace("-EQ", "")

def gf(d, *keys):
    """Get first non-empty value from multiple field names."""
    for k in keys:
        v = d.get(k)
        if v not in (None, "", 0, 0.0):
            return v
    return "–"

def safe_ratio(bq, sq):
    try:
        return round(int(bq) / int(sq), 2) if int(sq) > 0 else "∞"
    except Exception:
        return "–"

def safe_proxy(bq, sq):
    try:
        return int(bq) + int(sq)
    except Exception:
        return "–"

# ═══════════════════════════════════════════════════════
#  TELEGRAM
# ═══════════════════════════════════════════════════════
def send_telegram(text: str) -> None:
    try:
        r = requests.post(
            f"https://api.telegram.org/bot{TG_TOKEN}/sendMessage",
            data={"chat_id": TG_CHAT_ID, "text": text[:4096], "parse_mode": "HTML"},
            timeout=8
        )
        if r.status_code != 200:
            print(f"[TG WARN] {r.status_code}: {r.text[:80]}")
    except Exception as e:
        print(f"[TG ERROR] {e}")


def build_snap_msg(label: str) -> str:
    ts  = datetime.now().strftime("%H:%M:%S")
    dt  = datetime.now().strftime("%d %b %Y")
    ph  = get_phase(ts)

    with lock:
        data   = dict(latest)
        counts = dict(tick_count)
        hist   = {k: list(v) for k, v in ltp_history.items()}
        n_keys = len(all_keys)
        total  = sum(counts.values())

    lines = [
        f"<b>{label}</b>",
        f"<i>📅 {dt}  ⏰ {ts} IST  [{ph}]</i>",
        "<pre>" + "─"*56 + "</pre>",
        "<pre>SYM         IEP(open) PREV   GAP%  BUY_Q   SELL_Q  RATIO</pre>",
        "<pre>" + "─"*56 + "</pre>",
    ]

    iep_lines = ["\n<b>📌 IEP & LTP Status:</b>"]
    vol_lines = ["\n<b>📦 Volume (0=pre-open, >0=trades started):</b>"]

    for sym in SYMBOLS:
        s    = short(sym)
        d    = data.get(sym, {})
        iep  = d.get("open_price",       "–")
        prev = d.get("prev_close_price",  "–")
        chp  = gf(d, "chp", "change_perc")
        bq   = gf(d, "tot_buy_qty",  "totalbuyqty")
        sq   = gf(d, "tot_sell_qty", "totalsellqty")
        vol  = d.get("vol_traded_today", 0)
        tks  = counts.get(sym, 0)
        ratio= safe_ratio(bq, sq)

        # Formatted numbers
        try:    bq_fmt = f"{int(bq):,}"
        except: bq_fmt = str(bq)
        try:    sq_fmt = f"{int(sq):,}"
        except: sq_fmt = str(sq)

        lines.append(
            f"<pre>{s:<11}{str(iep):>9} {str(prev):>6} {str(chp):>6}% "
            f"{bq_fmt:>8} {sq_fmt:>8} {str(ratio):>5}</pre>"
        )

        # IEP analysis
        ltp_vals = hist.get(sym, [])
        moved    = "✅ MOVED" if len(set(ltp_vals)) > 1 else "🔴 FROZEN"
        try:
            iep_flag = "⚡ IEP≠prev" if float(iep) != float(prev) else "  IEP=prev"
        except Exception:
            iep_flag = "  ?"
        iep_lines.append(f"<pre>{s:<11} {iep_flag} | LTP {moved}</pre>")

        # Volume
        try:    vol_fmt = f"{int(vol):,}"
        except: vol_fmt = str(vol)
        vol_lines.append(f"<pre>{s:<11} vol={vol_fmt}</pre>")

    lines.append("<pre>" + "─"*56 + "</pre>")
    lines += iep_lines
    lines += vol_lines
    lines.append(f"\n<i>Total ticks: {total} | Unique fields: {n_keys}</i>")
    return "\n".join(lines)


# ═══════════════════════════════════════════════════════
#  CSV WRITERS
# ═══════════════════════════════════════════════════════
STANDARD_COLS = [
    "_ts", "_phase", "symbol",
    "ltp", "open_price", "prev_close_price",
    "ch", "chp", "tot_buy_qty", "tot_sell_qty",
    "bs_ratio", "proxy_vol",
    "vol_traded_today", "high_price", "low_price",
    "avg_trade_price", "lower_ckt", "upper_ckt",
    "last_traded_qty", "last_traded_time", "exch_feed_time",
    "type"
]

def write_raw(row: dict) -> None:
    global raw_writer
    if raw_writer is None:
        raw_writer = csv.DictWriter(RAW_FILE, fieldnames=STANDARD_COLS, extrasaction="ignore")
        raw_writer.writeheader()
    # Detect new fields
    new = set(row.keys()) - set(STANDARD_COLS)
    if new:
        alert = f"🆕 NEW FIELD @ {row.get('_ts','?')}: {new}"
        print(f"\n{'!'*60}\n{alert}\n{'!'*60}")
        threading.Thread(target=send_telegram, args=(f"<b>{alert}</b>",), daemon=True).start()
    raw_writer.writerow(row)
    RAW_FILE.flush()


def flush_per_second() -> None:
    global sec_writer
    with lock:
        keys_to_flush = sorted(set(one_sec_log.keys()) - written_sec_keys)
        snap = {k: dict(one_sec_log[k]) for k in keys_to_flush}

    for sec_key in keys_to_flush:
        for sym, d in snap[sec_key].items():
            bq = d.get("tot_buy_qty", 0)
            sq = d.get("tot_sell_qty", 0)
            row = {
                "_second":   sec_key,
                "_phase":    get_phase(sec_key),
                "symbol":    sym,
                "ticks":     tick_count.get(sym, 0),
                "bs_ratio":  safe_ratio(bq, sq),
                "proxy_vol": safe_proxy(bq, sq),
            }
            row.update(d)
            if sec_writer is None:
                sec_writer = csv.DictWriter(SEC_FILE, fieldnames=list(row.keys()), extrasaction="ignore")
                sec_writer.writeheader()
            sec_writer.writerow(row)
        with lock:
            written_sec_keys.add(sec_key)
    SEC_FILE.flush()


def write_phase_snapshot(label: str) -> None:
    global phase_writer
    ts = datetime.now().strftime("%H:%M:%S.%f")
    with lock:
        data   = dict(latest)
        counts = dict(tick_count)
        hist   = {k: list(v) for k, v in ltp_history.items()}

    for sym in SYMBOLS:
        d      = data.get(sym, {})
        bq     = d.get("tot_buy_qty", 0)
        sq     = d.get("tot_sell_qty", 0)
        ltp_v  = hist.get(sym, [])
        row = {
            "_ts":          ts,
            "_phase_label": label,
            "_phase":       get_phase(ts[:8]),
            "symbol":       sym,
            "tick_count":   counts.get(sym, 0),
            "ltp_moved":    1 if len(set(ltp_v)) > 1 else 0,
            "bs_ratio":     safe_ratio(bq, sq),
            "proxy_vol":    safe_proxy(bq, sq),
        }
        row.update(d)
        if phase_writer is None:
            phase_writer = csv.DictWriter(PHASE_FILE, fieldnames=list(row.keys()), extrasaction="ignore")
            phase_writer.writeheader()
        phase_writer.writerow(row)
    PHASE_FILE.flush()


# ═══════════════════════════════════════════════════════
#  EXCEL EXPORT  (V4's structured sheets)
# ═══════════════════════════════════════════════════════
def build_excel() -> None:
    if not EXCEL_OK:
        print("[EXCEL] Skipped — pip install openpyxl")
        return
    try:
        # Read raw CSV back and split by phase
        import csv as _csv
        phase_rows: dict = {p: [] for p in [
            "PRE_BASELINE","PHASE1_ORDER_COLLECTION",
            "PHASE2_MATCHING","PHASE3_BUFFER","REGULAR_MARKET"
        ]}
        with open(RAW_PATH, encoding="utf-8") as f:
            reader = _csv.DictReader(f)
            for row in reader:
                ph = row.get("_phase", "UNKNOWN")
                phase_rows.setdefault(ph, []).append(row)

        wb  = openpyxl.Workbook()
        wb.remove(wb.active)

        HDR_FILL  = PatternFill("solid", fgColor="1F4E79")
        FILLS = {
            "PRE_BASELINE":             PatternFill("solid", fgColor="F5F5F5"),
            "PHASE1_ORDER_COLLECTION":  PatternFill("solid", fgColor="E8F5E9"),
            "PHASE2_MATCHING":          PatternFill("solid", fgColor="FFF9C4"),
            "PHASE3_BUFFER":            PatternFill("solid", fgColor="FFE0B2"),
            "REGULAR_MARKET":           PatternFill("solid", fgColor="E3F2FD"),
        }

        DISPLAY_COLS = [
            ("_ts",              "Timestamp"),
            ("symbol",           "Symbol"),
            ("open_price",       "IEP (open_price)"),
            ("prev_close_price", "Prev Close"),
            ("ch",               "Gap ₹"),
            ("chp",              "Gap %"),
            ("ltp",              "LTP"),
            ("tot_buy_qty",      "Buy Qty"),
            ("tot_sell_qty",     "Sell Qty"),
            ("bs_ratio",         "B/S Ratio"),
            ("proxy_vol",        "Proxy Vol"),
            ("vol_traded_today", "Volume"),
            ("high_price",       "High"),
            ("low_price",        "Low"),
            ("lower_ckt",        "Lower CKT"),
            ("upper_ckt",        "Upper CKT"),
        ]

        for phase_key in ["PRE_BASELINE","PHASE1_ORDER_COLLECTION",
                          "PHASE2_MATCHING","PHASE3_BUFFER","REGULAR_MARKET"]:
            rows = phase_rows.get(phase_key, [])
            ws   = wb.create_sheet(title=phase_key[:20])

            # Header
            for ci, (_, hdr) in enumerate(DISPLAY_COLS, 1):
                c = ws.cell(row=1, column=ci, value=hdr)
                c.font      = Font(bold=True, color="FFFFFF")
                c.fill      = HDR_FILL
                c.alignment = Alignment(horizontal="center")

            fill = FILLS.get(phase_key, PatternFill())
            for ri, row in enumerate(rows, 2):
                for ci, (col, _) in enumerate(DISPLAY_COLS, 1):
                    ws.cell(row=ri, column=ci, value=row.get(col, "")).fill = fill

            ws.column_dimensions["A"].width = 18
            ws.column_dimensions["B"].width = 20
            for col_letter in ["C","D","E","F","G","H","I","J","K","L","M","N","O","P"]:
                ws.column_dimensions[col_letter].width = 14
            ws.freeze_panes = "A2"
            print(f"[EXCEL] {phase_key}: {len(rows)} rows")

        # SUMMARY SHEET — one row per stock
        ws_s = wb.create_sheet(title="SUMMARY", index=0)
        sum_hdrs = [
            "Symbol", "Final IEP", "Prev Close", "Gap ₹", "Gap %",
            "Phase1 Avg Buy", "Phase1 Avg Sell", "Phase1 B/S Ratio",
            "Phase1 Proxy Vol", "Phase1 Ticks",
            "Final LTP", "Normal Avg Volume",
            "IEP Moved?", "LTP Moved Phase1?"
        ]
        for ci, h in enumerate(sum_hdrs, 1):
            c = ws_s.cell(row=1, column=ci, value=h)
            c.font      = Font(bold=True, color="FFFFFF")
            c.fill      = HDR_FILL
            c.alignment = Alignment(horizontal="center")

        SUM_FILL = PatternFill("solid", fgColor="F3E5F5")
        with lock:
            last_data = dict(latest)
            hist_data = {k: list(v) for k, v in ltp_history.items()}

        for ri, sym in enumerate(SYMBOLS, 2):
            p1 = [r for r in phase_rows.get("PHASE1_ORDER_COLLECTION",[]) if r.get("symbol")==sym]
            nm = [r for r in phase_rows.get("REGULAR_MARKET",[])           if r.get("symbol")==sym]
            d  = last_data.get(sym, {})

            def avg(rows, key):
                vals = []
                for r in rows:
                    try: vals.append(float(r[key]))
                    except Exception: pass
                return round(sum(vals)/len(vals), 2) if vals else "–"

            prev       = d.get("prev_close_price", "–")
            final_iep  = d.get("open_price", "–")
            try:    gap_rs  = round(float(final_iep) - float(prev), 2)
            except: gap_rs  = "–"
            try:    gap_pct = round((float(gap_rs)/float(prev))*100, 3)
            except: gap_pct = "–"

            avg_bq   = avg(p1, "tot_buy_qty")
            avg_sq   = avg(p1, "tot_sell_qty")
            iep_vals = list(set(r.get("open_price","") for r in p1 if r.get("open_price") not in ("","–",None,"0","0.0")))
            ltp_vals = list(set(hist_data.get(sym, [])))

            row_data = [
                short(sym), final_iep, prev, gap_rs, gap_pct,
                avg_bq, avg_sq, safe_ratio(avg_bq, avg_sq),
                safe_proxy(avg_bq, avg_sq), len(p1),
                d.get("ltp","–"), avg(nm, "vol_traded_today"),
                "✅ YES" if len(iep_vals) > 1 else "🔴 NO",
                "✅ YES" if len(ltp_vals) > 1 else "🔴 NO",
            ]
            for ci, val in enumerate(row_data, 1):
                ws_s.cell(row=ri, column=ci, value=val).fill = SUM_FILL

        ws_s.column_dimensions["A"].width = 16
        for col_letter in ["B","C","D","E","F","G","H","I","J","K","L","M","N"]:
            ws_s.column_dimensions[col_letter].width = 20
        ws_s.freeze_panes = "A2"

        wb.save(XLS_PATH)
        print(f"[EXCEL] ✅ Saved: {XLS_PATH}")
    except Exception as e:
        print(f"[EXCEL ERROR] {e}")
        import traceback; traceback.print_exc()


# ═══════════════════════════════════════════════════════
#  SHUTDOWN
# ═══════════════════════════════════════════════════════
def shutdown(reason: str = "AUTO-STOP") -> None:
    if stop_event.is_set():
        return
    stop_event.set()

    ts = datetime.now().strftime("%H:%M:%S")
    print(f"\n{'='*62}\n⏹  {reason} @ {ts}\n{'='*62}")

    # Final Telegram
    send_telegram(build_snap_msg(f"✅ {reason} — Session Complete"))

    # Flush per-second
    flush_per_second()

    # Phase snapshot
    write_phase_snapshot("SESSION_END")

    # Close CSVs
    for f in [RAW_FILE, SEC_FILE, PHASE_FILE]:
        try: f.close()
        except: pass

    # Print final summary
    with lock:
        counts = dict(tick_count)
        keys   = sorted(all_keys)

    print("\nFINAL TICK COUNT:")
    for sym, cnt in counts.items():
        print(f"  {short(sym):<20} {cnt:>6} ticks")
    print(f"\nALL FIELDS ({len(keys)}): {keys}")
    print(f"\n✅ CSVs saved:")
    print(f"   {RAW_PATH}")
    print(f"   {SEC_PATH}")
    print(f"   {PHASE_PATH}")

    # Build Excel
    print("\nBuilding Excel...")
    build_excel()
    print(f"{'='*62}")


# ═══════════════════════════════════════════════════════
#  WEBSOCKET CALLBACKS
# ═══════════════════════════════════════════════════════
def onmessage(msg: dict) -> None:
    if stop_event.is_set():
        return

    ts    = datetime.now().strftime("%H:%M:%S.%f")[:-3]
    hms   = ts[:8]
    sym   = msg.get("symbol", msg.get("s", ""))
    phase = get_phase(hms)

    # Skip control messages and depth (confirmed useless)
    if msg.get("type") in ("cn","ful","sub","unsub","lit","cp","cr"):
        print(f"[{ts}] CTRL | {msg.get('type')}")
        return
    if not sym or msg.get("type") == "dp":
        return

    # ── Thread-safe state update ──────────────────────
    with lock:
        all_keys.update(msg.keys())
        latest.setdefault(sym, {}).update(msg)
        tick_count[sym] = tick_count.get(sym, 0) + 1
        tks = tick_count[sym]

        ltp_val = msg.get("ltp")
        if ltp_val:
            ltp_history.setdefault(sym, []).append(ltp_val)

        # Per-second log
        one_sec_log.setdefault(hms, {}).setdefault(sym, {}).update(msg)

    # ── Compute derived fields ────────────────────────
    iep      = msg.get("open_price", "")
    prev     = msg.get("prev_close_price", "")
    ltp      = msg.get("ltp", "–")
    chp      = msg.get("chp", "–")
    ch       = msg.get("ch",  "–")
    bq       = msg.get("tot_buy_qty",  "–")
    sq       = msg.get("tot_sell_qty", "–")
    bs_ratio = safe_ratio(bq, sq)
    proxy    = safe_proxy(bq, sq)

    try:    iep_diff = float(iep) != float(prev)
    except: iep_diff = False
    iep_flag = " ⚡IEP≠prev" if iep_diff else ""

    # ── ONE-LINE terminal (V2's clean style) ──────────
    print(
        f"[{ts}] {short(sym):<13} "
        f"IEP={str(iep):<8} LTP={str(ltp):<8} "
        f"BQ={str(bq):<10} SQ={str(sq):<10} "
        f"ratio={str(bs_ratio):<6} phase={phase} #{tks}"
        f"{iep_flag}"
    )

    # ── Write raw CSV ─────────────────────────────────
    row = {
        "_ts": ts, "_phase": phase, "symbol": sym,
        "ltp": ltp, "open_price": iep, "prev_close_price": prev,
        "ch": ch, "chp": chp,
        "tot_buy_qty": bq, "tot_sell_qty": sq,
        "bs_ratio": bs_ratio, "proxy_vol": proxy,
        "vol_traded_today":  msg.get("vol_traded_today", ""),
        "high_price":        msg.get("high_price", ""),
        "low_price":         msg.get("low_price", ""),
        "avg_trade_price":   msg.get("avg_trade_price", ""),
        "lower_ckt":         msg.get("lower_ckt", ""),
        "upper_ckt":         msg.get("upper_ckt", ""),
        "last_traded_qty":   msg.get("last_traded_qty", ""),
        "last_traded_time":  msg.get("last_traded_time", ""),
        "exch_feed_time":    msg.get("exch_feed_time", ""),
        "type":              msg.get("type", ""),
    }
    write_raw(row)

    # ── IEP instant alert (first move per stock) ──────
    if iep_diff and sym not in iep_alerted:
        iep_alerted.add(sym)
        threading.Thread(target=send_telegram, args=(
            f"⚡ <b>IEP FIRST CHANGE — {short(sym)}</b>\n"
            f"Time: {ts} IST\n"
            f"IEP: <b>{iep}</b>  |  Prev Close: {prev}\n"
            f"Gap: {ch} ({chp}%)\n"
            f"Phase: {phase}",
        ), daemon=True).start()

    # ── Scheduled Telegram snapshots ──────────────────
    for snap_time, label in SNAPSHOTS.items():
        if hms >= snap_time and snap_time not in sent_snaps:
            sent_snaps.add(snap_time)
            write_phase_snapshot(label)
            threading.Thread(
                target=send_telegram,
                args=(build_snap_msg(label),),
                daemon=True
            ).start()
            print(f"\n{'='*62}\n📬 TELEGRAM: {label}\n{'='*62}")


def onerror(msg) -> None:
    ts = datetime.now().strftime("%H:%M:%S")
    print(f"\n[❌ ERROR @ {ts}] {msg}")
    send_telegram(f"❌ <b>WS Error @ {ts}</b>\n{msg}")


def onclose(msg) -> None:
    print(f"\n[🔴 CLOSED] {msg}")
    shutdown("WS CLOSED")


def onopen() -> None:
    ts = datetime.now().strftime("%H:%M:%S")
    print(f"\n✅ WebSocket CONNECTED @ {ts}")
    print(f"📡 Subscribing SymbolUpdate only (DepthUpdate skipped)...")

    fyers.subscribe(symbols=SYMBOLS, data_type="SymbolUpdate")

    print(f"✅ Subscribed: {[short(s) for s in SYMBOLS]}")
    send_telegram(
        f"🔌 <b>Pre-Open Capture V5 CONNECTED</b>\n"
        f"<i>📅 {datetime.now().strftime('%d %b %Y')}  ⏰ {ts} IST</i>\n\n"
        f"<b>Symbols:</b> {[short(s) for s in SYMBOLS]}\n"
        f"<b>Window:</b> {START_TIME} → {STOP_TIME} (auto-stop)\n\n"
        f"<b>📁 Output files:</b>\n"
        f"  01_raw_ticks_{DATE_STR}.csv\n"
        f"  02_per_second_{DATE_STR}.csv\n"
        f"  03_phase_summary_{DATE_STR}.csv\n"
        f"  04_analysis_{DATE_STR}.xlsx\n\n"
        f"<b>📬 Snapshots at:</b>\n"
        f"  9:00 | 9:04 | 9:07 | 9:08 | 9:12 | 9:14 | 9:15 | 9:20"
    )
    fyers.keep_running()


# ═══════════════════════════════════════════════════════
#  BACKGROUND LOOP — 10s table + 60s flush + auto-stop
# ═══════════════════════════════════════════════════════
def background_loop() -> None:
    counter = 0
    while not stop_event.is_set():
        time.sleep(10)
        if stop_event.is_set():
            break
        counter += 1
        now   = datetime.now().strftime("%H:%M:%S")
        phase = get_phase(now)

        # ── AUTO-STOP ──────────────────────────────────
        if now >= STOP_TIME:
            print(f"\n⏰ AUTO-STOP triggered @ {now}")
            shutdown("AUTO-STOP")
            try: fyers.close_connection()
            except: pass
            os._exit(0)

        # ── FLUSH per-second CSV every 60s ─────────────
        if counter % 6 == 0:
            flush_per_second()
            with lock:
                flushed = len(written_sec_keys)
            print(f"[{now}] 💾 Per-second CSV flushed ({flushed} seconds logged)")

        # ── TERMINAL 10s SUMMARY TABLE ─────────────────
        with lock:
            data   = dict(latest)
            counts = dict(tick_count)
            total  = sum(counts.values())

        print(f"\n[{now}] ── {phase} ── 10s SUMMARY {'─'*35}")
        print(f"  {'SYM':<14}{'IEP':>9}{'LTP':>8}{'GAP%':>7}{'BUY_Q':>12}{'SELL_Q':>11}{'RATIO':>7}{'PROXY':>12}{'TICKS':>7}")
        print("  " + "─" * 87)
        for sym in SYMBOLS:
            d      = data.get(sym, {})
            iep    = d.get("open_price", "–")
            ltp    = d.get("ltp", "–")
            chp    = d.get("chp", d.get("change_perc", "–"))
            bq     = d.get("tot_buy_qty",  d.get("totalbuyqty",  "–"))
            sq     = d.get("tot_sell_qty", d.get("totalsellqty", "–"))
            tks    = counts.get(sym, 0)
            ratio  = safe_ratio(bq, sq)
            proxy  = safe_proxy(bq, sq)
            try:    bq_fmt = f"{int(bq):,}"
            except: bq_fmt = str(bq)
            try:    sq_fmt = f"{int(sq):,}"
            except: sq_fmt = str(sq)
            try:    pv_fmt = f"{int(proxy):,}"
            except: pv_fmt = str(proxy)
            try:    flag = "⚡" if float(iep) != float(d.get("prev_close_price",0)) else "  "
            except: flag = "  "
            print(f"  {flag}{short(sym):<12}{str(iep):>9}{str(ltp):>8}{str(chp):>7}%"
                  f"{bq_fmt:>12}{sq_fmt:>11}{str(ratio):>7}{pv_fmt:>12}{tks:>7}")
        print(f"  → Total ticks: {total} | Fields: {len(all_keys)}")


# ═══════════════════════════════════════════════════════
#  WAIT UNTIL START  (V2's feature — run anytime!)
# ═══════════════════════════════════════════════════════
def wait_until(target_hms: str) -> None:
    while True:
        now = datetime.now().strftime("%H:%M:%S")
        if now >= target_hms:
            print(f"\n✅ {target_hms} reached — connecting now...\n")
            return
        try:
            t1   = datetime.strptime(now,        "%H:%M:%S")
            t2   = datetime.strptime(target_hms, "%H:%M:%S")
            secs = int((t2 - t1).total_seconds())
            print(f"  ⏳ Waiting for {target_hms}... {secs}s remaining", end="\r")
        except: pass
        time.sleep(5)


# ═══════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════
if __name__ == "__main__":
    print("=" * 62)
    print("   NSE PRE-OPEN LIVE CAPTURE V5 — FINAL")
    print(f"   Date     : {DATE_STR}")
    print(f"   Window   : {START_TIME} → {STOP_TIME} (auto-stop)")
    print(f"   Symbols  : {[short(s) for s in SYMBOLS]}")
    print(f"   Telegram : Chat {TG_CHAT_ID}")
    print(f"   Outputs  : outputs/01_raw_ticks / 02_per_second / 03_phase_summary / 04_analysis")
    print("=" * 62)

    if not ACCESS_TOKEN:
        raise ValueError("❌ FYERS_ACCESS_TOKEN missing in .env")

    # ── Wait until 8:58 ───────────────────────────────
    wait_until(START_TIME)

    # ── Init WebSocket ────────────────────────────────
    fyers = data_ws.FyersDataSocket(
        access_token=FULL_TOKEN,
        write_to_file=False,
        log_path="",
        litemode=False,
        reconnect=True,
        reconnect_retry=3,
        on_connect=onopen,
        on_message=onmessage,
        on_error=onerror,
        on_close=onclose,
    )

    # ── Start background thread ───────────────────────
    threading.Thread(target=background_loop, daemon=True).start()

    # ── Connect ───────────────────────────────────────
    try:
        fyers.connect()
    except KeyboardInterrupt:
        print("\n⏹ Ctrl+C — saving all files...")
        shutdown("MANUAL CTRL+C")